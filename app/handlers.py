from aiogram import Router, F

from aiogram.filters import CommandStart
from aiogram.types import Message, CallbackQuery, ReplyKeyboardRemove, FSInputFile

from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext

from datetime import datetime
from datetime import time
from haversine import haversine, Unit
from apscheduler.schedulers.asyncio import AsyncIOScheduler

import app.keyboards as kb
import database.requests as bd
import sites.purse as purse
import dateparser
import openpyxl


class Student(StatesGroup):
    institute = State()
    education = State()
    fuclty = State()
    group = State()
    name = State()
    schedule = State()


router = Router()

start_text = """
Привет!😁✌️
Здесь ты можешь:
- отметиться в списке присутвующих на занятии
- получить расписание на неделю
- выгрузить список присутствующих за этот учебный год
"""
info_text = """Для навигации используй кнопки\n
📌<b>"Авторизоваться"</b> - чтобы найти себя в базе
📌<b>"Информация"</b> - чтобы вывести инфу о себе
📌<b>"Отметиться"</b> - чтобы отметиться на паре
учти, что отмечаться можно за 20 минут до начала занятия, а также проверяется геолокация😉 
📌<b>"Расписание"</b> - чтобы посмотреть расписание на текущую неделю
📌<b>"Списки присутствующих"</b> - чтобы создался файл со списками присутствующих"""

chat_ids = [1871402519]


async def clear_schedule(state: FSMContext):
    print("расписание отчистилось")
    await state.update_data(schedule=None)


@router.message(CommandStart())
async def command_start(message: Message, state: FSMContext) -> None:
    if message.from_user.id not in chat_ids:
        chat_ids.append(message.from_user.id)
    print(chat_ids)
    scheduler = AsyncIOScheduler()
    scheduler.add_job(clear_schedule, 'cron', minute=1, hour=0, day_of_week='mon', args=[state])
    scheduler.start()
    await message.delete()
    await message.answer(start_text, reply_markup=kb.first_kb)
    await message.answer(info_text, parse_mode="html")
    await message.answer_sticker("CAACAgIAAxkBAAEMARFmLVqDWXJu6zTfHeHGH9Ug8Eqx6wACyUAAAr5sEEgMQFhAO8zizTQE")


# @router.message(Command('help'))
# async def command_help(message: Message) -> None:
#     await message.answer(help_text, parse_mode='HTML')


@router.message(F.text == 'Авторизоваться')
async def find_group(message: Message, state: FSMContext) -> None:
    await message.delete()
    sent_message = await message.answer(text="Начинается авторизация", reply_markup=ReplyKeyboardRemove())
    await sent_message.delete()
    state_data = await state.get_data()
    schedule = state_data.get('schedule')
    if schedule is not None and not (schedule == "На эту неделю занятия не поставлены 😘"):
        # await message.answer("")
        print("deleting")
        student_id = state_data.get('student_id')
        group_id = state_data.get('study_group_id')

        for day in schedule:
            times = []
            if not datetime.now().date() == dateparser.parse(day['date']).date():
                print("Даты не совпадают")
            else:
                for lesson in day['lessons']:
                    start_h = int(lesson['time_start'].split(":")[0])
                    start_m = int(lesson['time_start'].split(':')[1])
                    end_h = int(lesson['time_end'].split(':')[0])
                    end_m = int(lesson['time_end'].split(':')[1])
                    spot_lesson = lesson['spot']
                    name_lesson = lesson['name_lesson']
                    times.append([time(start_h, start_m), time(end_h, end_m), spot_lesson, name_lesson])
                # times.append([time(1, 00), time(3, 40), "Главный учебный корпус", "Home"])
                # times.append([time(14, 00), time(16, 40), "Главный учебный корпус", "Discord"])

                # print(times)
                for start, end, spot, name in times:
                    print(f"{time(start.hour - 1, start.minute + 40)} - {end} - {spot}")
                    start_time = time(start.hour - 1, start.minute + 40)
                    now_time = time(datetime.now().hour, datetime.now().minute)
                    if start_time <= now_time <= end:
                        bd.delete_visiting(lesson=name,
                                           group_id=group_id,
                                           student_id=student_id,
                                           year=datetime.now().year,
                                           month=datetime.now().month,
                                           day=datetime.now().day)


    await state.update_data(institute_id=None, fuclty_id=None, study_group_id=None,
                                student_id=None, education=None, schedule=None)
    await message.answer("Выбери сециальность", reply_markup=kb.get_education_kb())
        # await message.delete()


@router.message(F.text == 'Отметиться')
async def get_location(message: Message, state: FSMContext):
    state_data = await state.get_data()
    group_id = state_data.get('study_group_id')
    url = bd.get_schedule_url(group_id)
    if group_id is None:
        await message.answer("Так ты сначала авторизуйся 😡")
        await message.answer_sticker("CAACAgQAAxkBAAEMC1dmNK3bWk21XDnN-lEU2XG5EtbuSwACjwcAAvnhgFEO9a_zf5fHtDQE")
    elif url is None:
        await message.answer("На эту неделю занятия не поставлены 😘")
    else:
        sent_message = await message.answer("Подождите, данные загружаются...")
        schedule_state = state_data.get('schedule')
        if schedule_state is None:
            schedule = purse.get_schedule(url)
            await state.update_data(schedule=schedule)
        state_data = await state.get_data()
        schedule = state_data.get('schedule')
        await sent_message.delete()

        if schedule == "На эту неделю занятия не поставлены 😘":
            await message.answer(schedule)
        else:
            flag = True
            for day in schedule:
                times = []
                if not datetime.now().date() == dateparser.parse(day['date']).date():
                    print("Даты не совпадают")
                else:
                    print("Даты совпадают")
                    for lesson in day['lessons']:

                        start_h = int(lesson['time_start'].split(":")[0])
                        start_m = int(lesson['time_start'].split(':')[1])
                        end_h = int(lesson['time_end'].split(':')[0])
                        end_m = int(lesson['time_end'].split(':')[1])
                        spot_lesson = lesson['spot']
                        name_lesson = lesson['name_lesson']

                        times.append([time(start_h, start_m), time(end_h, end_m), spot_lesson, name_lesson])

                    # times.append([time(14, 00), time(16, 40), "Главный учебный корпус", "Discord"])

                    for start, end, spot, name in times:
                        #print(f"{time(start.hour - 1, start.minute + 40)} - {end} - {spot}")
                        start_time = time(start.hour - 1, start.minute + 40)
                        now_time = time(datetime.now().hour, datetime.now().minute)
                        if start_time <= now_time <= end:
                            flag = False
                            #print(f"now: {time(start.hour - 1, start.minute + 40)} - {end} - {spot}")
                            mess_text = ("Вы собираетесь отметиться на паре \n\n<b>Предмет:</b> " + name +
                                    "\n\n<b>Время:</b> " + start.strftime("%H:%M") + "-" + end.strftime("%H:%M") +
                                    "\n\n<b>Место проведения:</b> " + spot)
                            await message.answer(mess_text, parse_mode="html")
                            await message.answer("Пожалуйста, отправьте ваше местоположение.",
                                                 reply_markup=kb.geo_kb)

            if flag:
                await message.answer("В данный момент ты ни на что не можешь отметиться",
                                     reply_markup=kb.cancel_kb)
                sticker_id = "CAACAgIAAxkBAAEMETpmOojZfGmJtZsVqrGTqV1AB1tangACORYAAjL-yEpvoPRnuqpSqTUE"
                await message.answer_sticker(sticker_id)


@router.message(F.location)
async def handle_location(message: Message, state: FSMContext):
    latitude = message.location.latitude
    longitude = message.location.longitude
    state_data = await state.get_data()
    schedule = state_data.get('schedule')
    student_id = state_data.get('student_id')
    group_id = state_data.get('study_group_id')
    if not schedule == "На эту неделю занятия не поставлены 😘":
        for day in schedule:
            times = []
            if not datetime.now().date() == dateparser.parse(day['date']).date():
                print("Даты не совпадают")
            else:
                for lesson in day['lessons']:
                    start_h = int(lesson['time_start'].split(":")[0])
                    start_m = int(lesson['time_start'].split(':')[1])
                    end_h = int(lesson['time_end'].split(':')[0])
                    end_m = int(lesson['time_end'].split(':')[1])
                    spot_lesson = lesson['spot']
                    name_lesson = lesson['name_lesson']
                    times.append([time(start_h, start_m), time(end_h, end_m), spot_lesson, name_lesson])
                # times.append([time(1, 00), time(3, 40), "Главный учебный корпус", "Home"])
                # times.append([time(14, 00), time(16, 40), "Главный учебный корпус", "Discord"])

                # print(times)
                for start, end, spot, name in times:
                    print(f"{time(start.hour - 1, start.minute + 40)} - {end} - {spot}")
                    start_time = time(start.hour - 1, start.minute + 40)
                    now_time = time(datetime.now().hour, datetime.now().minute)
                    if start_time <= now_time <= end:
                        geo = bd.get_geo(spot)
                        spot_geo = (float(geo[0]), float(geo[1]))
                        student_geo = (latitude, longitude)
                        distance = haversine(spot_geo, student_geo, unit=Unit.METERS)
                        print(distance)

                        if distance > 200:
                            mess = ("Ты НЕ смог отметиться на паре\n\n" +
                                    "Чтобы отметиться нужно находиться до 200 метров от корпуса\n\n" +
                                    "Твое расстояние до  учебного корпуса: ~" + str(round(distance)) + " метров")
                            await message.answer(mess, reply_markup=kb.geo_kb)
                        else:
                            bd.insert_lesson(lesson=name,
                                             group_id=group_id,
                                             student_id=student_id,
                                             year=datetime.now().year,
                                             month=datetime.now().month,
                                             day=datetime.now().day)


@router.message(F.text == 'Списки присутствующих')
async def send_file(message: Message, state: FSMContext) -> None:
    state_data = await state.get_data()
    group_id = state_data.get('study_group_id')
    # await message.answer(f'Время сейчас: {datetime.now()}')
    if group_id is None:
        await message.answer("Так ты сначала авторизуйся 😡")
        await message.answer_sticker("CAACAgQAAxkBAAEMC1dmNK3bWk21XDnN-lEU2XG5EtbuSwACjwcAAvnhgFEO9a_zf5fHtDQE")
    else:
        sent_mess = await message.answer("Придется подождать какоето время, лучше никуда не нажимай")

        book = openpyxl.Workbook()
        book.remove(book.active)

        lessons = bd.get_lessons(group_id)
        print(lessons)
        if lessons:
            for lesson in lessons:
                sheet = book.create_sheet(lesson)
                if datetime.now().date().month > 6:
                    year = str(datetime.now().date().year)
                else:
                    year = str(datetime.now().date().year - 1)
                day_start = year + "-8-30"
                dates = bd.get_lesson_dates(lesson=lesson, group_id=group_id, date_start=day_start)
                # sheet.append(dates)
                rows = [dates] + (bd.get_student_list(dates=dates[1:], group_id=group_id, lesson=lesson))
                for row in rows:
                    sheet.append(row)

                for col in sheet.columns:
                    max_length = 0
                    column = col[0].column_letter  # Получаем буквенное обозначение столбца (A, B, C, ...)
                    for cell in col:
                        try:
                            # Вычисляем максимальную длину значения в столбце
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    # Устанавливаем ширину столбца так, чтобы вмещалась самая длинная строка
                    adjusted_width = (max_length + 2) * 1.2
                    sheet.column_dimensions[column].width = adjusted_width

                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                    for cell in row:
                        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

                for row in sheet.iter_rows(min_row=1, max_row=1):
                    for cell in row:
                        cell.font = openpyxl.styles.Font(bold=True)
        else:
            book.create_sheet("данных пока нет")

        book.save("Посещаемость.xlsx")

        await sent_mess.delete()
        await message.answer("В файле списки присутствующих в твоей группе за текущий учебный год",
                             reply_markup=kb.first_kb)
        file_path = "D:\Another\pyCharm projects\Tg bot\Посещаемость.xlsx"
        await message.answer_document(document=FSInputFile(path=file_path))


@router.message(F.text == 'Информация')
async def print_info(message: Message, state: FSMContext) -> None:
    # await message.delete()
    state_data = await state.get_data()

    institute = "Неизвестно"
    institute_id = state_data.get('institute_id')
    if institute_id is not None:
        institute = bd.get_institute_from_id(institute_id)

    fuclty = "Неизвестно"
    fuclty_id = state_data.get('fuclty_id')
    if fuclty_id is not None:
        fuclty = bd.get_fuclty_from_id(fuclty_id)

    group = "Неизвестно"
    study_group_id = state_data.get('study_group_id')
    if study_group_id is not None:
        group = bd.get_study_group_from_id(study_group_id)

    student = "Неизвестно"
    student_id = state_data.get('student_id')
    if student_id is not None:
        student = bd.get_student_from_id(student_id)

    education = "Неизвестно"
    education_from_state = state_data.get('education')
    if education_from_state is not None:
        education = education_from_state

    print(f"institute: {institute_id}")
    print(f"fuclty: {fuclty_id}")
    print(f"group: {study_group_id}")
    print(f"student: {student_id}")
    information = ("""Информация про тебя:\n\n<b>Институт:</b>\n""" + institute
                   + "\n\n<b>Направление:</b>\n" + fuclty + "\n\n<b>Группа:</b>\n" + group
                   + "\n\n<b>Уровень образования:</b>\n" + education + "\n\nФИО:\n" + student)
    await message.answer(information, parse_mode="html")
    # await message.delete()


@router.message(F.text == 'Расписание')
async def show_schedule(message: Message, state: FSMContext):
    state_data = await state.get_data()
    group_id = state_data.get('study_group_id')
    url = bd.get_schedule_url(group_id)
    print(url)
    if group_id is None:
        await message.answer("Так ты сначала авторизуйся 😡")
        await message.answer_sticker("CAACAgQAAxkBAAEMC1dmNK3bWk21XDnN-lEU2XG5EtbuSwACjwcAAvnhgFEO9a_zf5fHtDQE")
    elif url is None:
        await message.answer("На эту неделю занятия не поставлены 😘")
    else:
        sent_message = await message.answer("Подождите, данные загружаются...")
        schedule_state = state_data.get('schedule')
        if schedule_state is None:
            schedule = purse.get_schedule(url)
            await state.update_data(schedule=schedule)
        state_data = await state.get_data()
        schedule = state_data.get('schedule')
        if schedule == "На эту неделю занятия не поставлены 😘":
            await message.answer(schedule)
        else:
            texts = []
            for day in schedule:
                text = "<b>" + day['date'] + "</b>\n\n\t"
                for lesson in day['lessons']:
                    text += lesson['time_start'] + "-" + lesson['time_end'] + "\n\t"
                    text += lesson['name_lesson'] + "\n\t"
                    text += lesson['lesson_type'] + "\n\t"
                    if not lesson['teacher'] == "":
                        text += lesson['teacher'] + "\n\t"
                    text += lesson['spot'] + ", ауд. " + lesson['auditory'] + "\n\n"
                texts.append(text)
            await sent_message.delete()
            for elem in texts:
                await message.answer(elem, parse_mode="html")


@router.message(F.text == '✖️Отмена')
async def undo_button(message: Message, state: FSMContext):
    await message.delete()
    await message.answer(info_text, reply_markup=kb.first_kb, parse_mode="html")
    await message.answer_sticker("CAACAgIAAxkBAAEMARVmLVqeD2IT8158d77O5mFlpgWr0wACyUAAAr5sEEgMQFhAO8zizTQE")


@router.message()
async def echo_handler(message: Message) -> None:
    await message.answer(text="Я тебя не понимаю...")
    await message.answer_sticker("CAACAgIAAxkBAAEMAQ5mLVcL7_veDeDVSOveH12ChgpjlAACGUIAAnn-6UgGEvxtyY9dhTQE")


@router.callback_query(lambda query: query.data.startswith("back_education"))
async def education(callback: CallbackQuery, state: FSMContext):
    await state.update_data(institute_id=None, fuclty_id=None, study_group_id=None,
                            student_id=None, education=None, schedule=None)
    await callback.message.delete()
    await callback.message.answer("Выбери сециальность", reply_markup=kb.get_education_kb())
    await callback.answer()


@router.callback_query(lambda query: query.data.startswith("education_"))
async def institute(callback: CallbackQuery, state: FSMContext):
    education = callback.data.split('_')[1]
    if education == "back":
        state_data = await state.get_data()
        education = state_data.get('education')
        await state.update_data(institute_id=None)
        print(callback.data)
        print(education)
    else:
        await state.update_data(education=education)
        print("succesfuly)))")
    await callback.message.delete()
    await callback.message.answer("Выбери свой институт",
                                  reply_markup=kb.get_institute_kb(education, 1),
                                  parse_mode="html")
    await callback.answer()


@router.callback_query(lambda query: query.data.startswith("institute_page_"))
async def institute_page(callback: CallbackQuery, state: FSMContext):
    page = int(callback.data.split('_')[2])
    await callback.message.delete()
    state_data = await state.get_data()
    education_state = state_data.get('education')
    await callback.message.answer("Выбери свой институт",
                                  reply_markup=kb.get_institute_kb(education_state, page),
                                  parse_mode="html")
    await callback.answer()


@router.callback_query(lambda query: query.data.startswith("id_institute_"))
async def fuclty(callback: CallbackQuery, state: FSMContext):
    institute_id = callback.data.split('_')[2]
    if institute_id == "back":
        state_data = await state.get_data()
        institute_id_state = state_data.get('institute_id')
        education_state = state_data.get('education')
        await state.update_data(fuclty_id=None)
        # print(callback.data)
        # print(education)
    else:
        institute_id = int(institute_id)
        await state.update_data(institute_id=institute_id)
        state_data = await state.get_data()
        institute_id_state = state_data.get('institute_id')
        education_state = state_data.get('education')
    # print("*&^" * 10)
    # print(f"institute:{institute_id_state}")
    await callback.message.delete()
    await callback.message.answer("""Выбери свое направление""",
                                  reply_markup=kb.get_fuclty_kb(institute_id_state, education_state, 1),
                                  parse_mode="html")
    await callback.answer()


@router.callback_query(lambda query: query.data.startswith("fuclty_page_"))
async def fuclty_page(callback: CallbackQuery, state: FSMContext):
    page = int(callback.data.split('_')[2])
    await callback.message.delete()
    state_data = await state.get_data()
    institute_id_state = state_data.get('institute_id')
    education_state = state_data.get('education')
    await callback.message.answer("Выбери свое направление",
                                  reply_markup=kb.get_fuclty_kb(institute_id_state, education_state, page),
                                  parse_mode="html")
    await callback.answer()


@router.callback_query(lambda query: query.data.startswith('id_fuclty_'))
async def group(callback: CallbackQuery, state: FSMContext):
    fuclty_id = callback.data.split('_')[2]
    if fuclty_id == "back":
        state_data = await state.get_data()
        fuclty_id_state = state_data.get('fuclty_id')
        education_state = state_data.get('education')
        await state.update_data(study_group_id=None)
        # print(callback.data)
        # print(education)
    else:
        fuclty_id = int(fuclty_id)
        await state.update_data(fuclty_id=fuclty_id)
        state_data = await state.get_data()
        fuclty_id_state = state_data.get('fuclty_id')
        education_state = state_data.get('education')
    # print("*&^" * 10)
    # print(f"institute:{fuclty_id_state}")
    # print(f"institute:{fuclty_id_state}")
    await callback.message.delete()
    await callback.message.answer("""Выбери свою группу\n1 курс:""",
                                  reply_markup=kb.get_group_kb(fuclty_id_state, 1, education_state, 1))
    await callback.answer()


@router.callback_query(lambda query: query.data.startswith("course_"))
async def group_course(callback: CallbackQuery, state: FSMContext):
    course = int(callback.data.split('_')[1])
    await callback.message.delete()
    state_data = await state.get_data()
    fuclty_id_state = state_data.get('fuclty_id')
    education_state = state_data.get('education')
    # print("*&^" * 10)
    # print(f"institute:{institute_id_state}")
    await callback.message.answer("Выбери свою группу\n" + str(course) + " курс:",
                                  reply_markup=kb.get_group_kb(fuclty_id_state, course, education_state, 1))
    await callback.answer()


@router.callback_query(lambda query: query.data.startswith("study_group_page_"))
async def group_page(callback: CallbackQuery, state: FSMContext):
    page = int(callback.data.split('_')[3])
    course = int(callback.data.split('_')[4])
    await callback.message.delete()
    state_data = await state.get_data()
    fuclty_id_state = state_data.get('fuclty_id')
    education_state = state_data.get('education')
    await callback.message.answer("Выбери свою группу\n" + str(course) + " курс:",
                                  reply_markup=kb.get_group_kb(fuclty_id_state, course, education_state, page),
                                  parse_mode="html")
    await callback.answer()


@router.callback_query(lambda query: query.data.startswith('id_study_group_'))
async def student(callback: CallbackQuery, state: FSMContext):
    study_group_id = int(callback.data.split('_')[3])
    await state.update_data(study_group_id=study_group_id)
    state_data = await state.get_data()
    study_group_id_state = state_data.get('study_group_id')
    # print("*&^" * 10)
    # print(f"institute:{fuclty_id_state}")
    await callback.message.delete()
    await callback.message.answer("""Выбери свою фамилию""",
                                  reply_markup=kb.get_student_kb(study_group_id_state, 1))
    await callback.answer()


@router.callback_query(lambda query: query.data.startswith("student_page_"))
async def student_page(callback: CallbackQuery, state: FSMContext):
    page = int(callback.data.split('_')[2])
    await callback.message.delete()
    state_data = await state.get_data()
    study_group_id_state = state_data.get('study_group_id')
    await callback.message.answer("Выбери свою фамилию",
                                  reply_markup=kb.get_student_kb(study_group_id_state, page),
                                  parse_mode="html")
    await callback.answer()


@router.callback_query(lambda query: query.data.startswith('id_student'))
async def student_done(callback: CallbackQuery, state: FSMContext):
    student_id = int(callback.data.split('_')[2])
    await state.update_data(student_id=student_id)
    await callback.message.delete()

    state_data = await state.get_data()

    institute = "Не известно"
    institute_id = state_data.get('institute_id')
    if institute_id is not None:
        institute = bd.get_institute_from_id(institute_id)

    fuclty = "Не известно"
    fuclty_id = state_data.get('fuclty_id')
    if fuclty_id is not None:
        fuclty = bd.get_fuclty_from_id(fuclty_id)

    group = "Не известно"
    study_group_id = state_data.get('study_group_id')
    if study_group_id is not None:
        group = bd.get_study_group_from_id(study_group_id)

    student = "Не известно"
    student_id = state_data.get('student_id')
    if student_id is not None:
        student = bd.get_student_from_id(student_id)

    education = "Не известно"
    education_from_state = state_data.get('education')
    if education_from_state is not None:
        education = education_from_state

    print(f"institute: {institute_id}")
    print(f"fuclty: {fuclty_id}")
    print(f"group: {study_group_id}")
    print(f"student: {student_id}")
    information = ("""Урааа, ты авторизовался!!!\nПроверь на всякий инфу про себя:\n\n<b>Институт:</b>\n""" + institute
                   + "\n\n<b>Направление:</b>\n" + fuclty + "\n\n<b>Группа:</b>\n" + group
                   + "\n\n<b>Уровень образования:</b>\n" + education + "\n\nФИО:\n" + student)
    await callback.message.answer(information, parse_mode="html")
    await callback.message.answer_sticker("CAACAgIAAxkBAAEMAw1mLoQn_rCjO8Xari_DzfzIriS1xgACjRYAAqBhwEsofH0_ZuoMbDQE")
    await callback.message.answer(info_text,reply_markup=kb.first_kb, parse_mode="html")
    await callback.message.answer_sticker("CAACAgIAAxkBAAEMARFmLVqDWXJu6zTfHeHGH9Ug8Eqx6wACyUAAAr5sEEgMQFhAO8zizTQE")
    await callback.answer()



@router.callback_query(lambda query: query.data.startswith('undo'))
async def cmd_undo(callback: CallbackQuery, state: FSMContext):
    await callback.message.delete()
    await state.update_data(institute_id=None, fuclty_id=None, study_group_id=None,
                            student_id=None, education=None, schedule=None)
    await callback.message.answer("Ты нажал кнопку отмены, твоя авторизация сбросилась!")
    await callback.message.answer_sticker("CAACAgQAAxkBAAEMAwdmLoCk8K-i0bZxUXGZ1xBI7tUVWgACjwcAAvnhgFEO9a_zf5fHtDQE")
    await callback.message.answer(info_text, reply_markup=kb.first_kb, parse_mode="html")
    await callback.message.answer_sticker("CAACAgIAAxkBAAEMARVmLVqeD2IT8158d77O5mFlpgWr0wACyUAAAr5sEEgMQFhAO8zizTQE")
    await callback.answer()
